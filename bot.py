import logging
import sqlite3
import os
from datetime import datetime
from collections import defaultdict
from io import BytesIO

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ApplicationBuilder, ContextTypes, MessageHandler, CallbackQueryHandler, filters
from PIL import Image, ImageDraw, ImageFont
from openpyxl.styles import PatternFill, Alignment

# 引入配置文件的名称处理逻辑
import config
import emoji


# 在文件最顶部添加
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
import pandas as pd # 新增


# ================= 配置区域 =================
BOT_TOKEN = '8450289917:AAHxeLh5Lkw9tECk-cV-hUptePq7S6smEQ0' 
DB_FILE = 'sanguosha.db'
FONT_PATH = 'simhei.ttf'  # 必须存在该字体文件
FONT_SIZE = 24
TRUNCATE_PASSWORD = 'qikenggouka'

logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)

# ================= 数据库操作 =================
def init_db():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS matches
                 (id INTEGER PRIMARY KEY AUTOINCREMENT, timestamp TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS details
                 (id INTEGER PRIMARY KEY AUTOINCREMENT, 
                  match_id INTEGER, 
                  player TEXT, 
                  main_gen TEXT, 
                  vice_gen TEXT, 
                  is_winner INTEGER,
                  FOREIGN KEY(match_id) REFERENCES matches(id))''')
    conn.commit()
    conn.close()

def get_db_connection():
    return sqlite3.connect(DB_FILE)



# ================= 邮件发送工具 =================

def generate_excel_bytes():
    """生成 Excel：去掉副标题行，去掉player行，手动合并局号表头，高亮胜者"""
    conn = get_db_connection()
    # 1. 读取原始数据
    query = """
    SELECT d.player, d.match_id, d.main_gen, d.vice_gen, d.is_winner 
    FROM details d
    ORDER BY d.match_id ASC
    """
    df = pd.read_sql_query(query, conn)
    conn.close()

    if df.empty:
        return None

    # 2. 构建胜负字典
    win_map = {}
    for _, row in df.iterrows():
        win_map[(row['player'], row['match_id'])] = bool(row['is_winner'])

    # 3. 数据透视
    pivot_df = df.pivot(index='player', columns='match_id', values=['main_gen', 'vice_gen'])
    
    # 4. 调整列顺序：让同一局的主副将挨在一起 (Match 1 Main, Match 1 Vice, Match 2...)
    pivot_df.columns = pivot_df.columns.swaplevel(0, 1)
    pivot_df.sort_index(axis=1, level=0, inplace=True)
    
    # 5. 准备辅助数据
    # 获取排序后的所有局号 (用于画表头)
    unique_match_ids = sorted(df['match_id'].unique())
    
    # 映射每一列对应的 match_id (用于填色)
    # pivot_df 的列现在是 [(1, main), (1, vice), (2, main)...]
    col_to_match_id = [col[0] for col in pivot_df.columns]

    # === 关键点 A: 删除索引名称 ===
    pivot_df.index.name = None

    # 6. 写入 Excel
    output = BytesIO()
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    center_align = Alignment(horizontal='center', vertical='center')

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # === 关键点 B: header=False, startrow=1 ===
        # header=False: 不写 "主将/副将" 这一行，也不写 pivot 自动生成的复杂表头
        # startrow=1: 数据从 Excel 的第 2 行开始写 (留出第 1 行给我们手动写局号)
        pivot_df.to_excel(writer, sheet_name='对局记录', header=False, startrow=1)
        
        workbook = writer.book
        worksheet = writer.sheets['对局记录']
        
        # 7. 手动绘制第一行表头 (第 X 局) 并合并单元格
        # Excel 列: A列是名字, B列开始是数据
        current_col = 2 
        
        for mid in unique_match_ids:
            # 写入 "第 X 局"
            cell = worksheet.cell(row=1, column=current_col)
            cell.value = f"第 {mid} 局"
            cell.alignment = center_align
            
            # 合并单元格 (覆盖主将和副将两列)
            # 例如: Merge B1:C1
            worksheet.merge_cells(start_row=1, end_row=1, 
                                  start_column=current_col, end_column=current_col + 1)
            
            current_col += 2

        # 8. 高亮胜者 (数据从第2行开始)
        start_row = 3 
        
        for i, player in enumerate(pivot_df.index):
            current_row = start_row + i
            
            for col_idx, match_id in enumerate(col_to_match_id):
                if win_map.get((player, match_id)):
                    # A列是玩家(1)，数据从B列(2)开始
                    excel_col = col_idx + 2
                    try:
                        cell = worksheet.cell(row=current_row, column=excel_col)
                        cell.fill = yellow_fill
                    except:
                        pass

    output.seek(0)
    return output

def send_email_with_excel(to_addr, excel_bytes):
    """发送带附件的邮件"""
    msg = MIMEMultipart()
    msg['From'] = config.SENDER_EMAIL
    msg['To'] = to_addr
    msg['Subject'] = "📊 蒸蒸日上 - 三国杀国战对局记录"

    body = "附件为您申请的对局记录 Excel 表格，请查收。\n\n——来自 Telegram Bot"
    msg.attach(MIMEText(body, 'plain'))

    # 添加附件
    part = MIMEBase('application', 'octet-stream')
    part.set_payload(excel_bytes.read())
    encoders.encode_base64(part)
    filename = f"sanguosha_records_{datetime.now().strftime('%Y%m%d')}.xlsx"
    part.add_header('Content-Disposition', f"attachment; filename= {filename}")
    msg.attach(part)

    # 发送
    try:
        server = smtplib.SMTP(config.SMTP_SERVER, config.SMTP_PORT)
        server.starttls()
        server.login(config.SENDER_EMAIL, config.SENDER_PASSWORD)
        text = msg.as_string()
        server.sendmail(config.SENDER_EMAIL, to_addr, text)
        server.quit()
        return True, "发送成功"
    except Exception as e:
        return False, str(e)
    



def draw_excel_style_image(date_str, matches_data, players_set):
    """
    matches_data结构: [(match_id, time_str, details_list), ...]
    details_list: [(player, main, vice, is_winner), ...]
    players_set: 当天所有参与过的玩家名字集合
    """
    
    # --- 1. 参数配置 ---
    # 颜色
    COLOR_BG = (255, 255, 255)       # 白底
    COLOR_GRID = (200, 200, 200)     # 灰线
    COLOR_TEXT = (0, 0, 0)           # 黑字
    COLOR_WIN_BG = (255, 255, 0)     # 胜者高亮(黄) - 参考你的截图
    
    # 尺寸
    CELL_W_NAME = 150   # 名字列宽
    CELL_W_GAME = 220   # 游戏列宽 (容纳两个武将名)
    ROW_H = 50          # 行高
    HEADER_H = 60       # 顶部日期栏高度
    
    font_size = 24
    try:
        font = ImageFont.truetype(FONT_PATH, font_size)
        font_bold = ImageFont.truetype(FONT_PATH, font_size + 4) # 标题稍大
    except:
        font = ImageFont.load_default()
        font_bold = ImageFont.load_default()

    # --- 2. 数据准备 ---
    sorted_players = sorted(list(players_set)) # 玩家按字母/拼音排序，固定行顺序
    match_ids = [m[0] for m in matches_data]   # 局号列表
    
    # 构建快速查找字典: data_map[match_id][player_name] = {'main':..., 'vice':..., 'win':...}
    data_map = defaultdict(lambda: defaultdict(dict))
    for mid, _, details in matches_data:
        for p, m, v, w in details:
            data_map[mid][p] = {'main': m, 'vice': v, 'win': w}

    # --- 3. 计算画布尺寸 ---
    cols = len(matches_data)
    rows = len(sorted_players)
    
    img_width = CELL_W_NAME + cols * CELL_W_GAME + 1 # +1 为了画最右边的线
    img_height = HEADER_H + ROW_H + rows * ROW_H + 1 # 日期头 + 局号头 + 玩家行
    
    image = Image.new('RGB', (img_width, img_height), COLOR_BG)
    draw = ImageDraw.Draw(image)

    # --- 4. 绘制函数辅助 ---
    def draw_cell_text(x, y, w, h, text, f=font, bg=None):
        # 填充背景
        if bg:
            draw.rectangle([x, y, x+w, y+h], fill=bg)
        # 画边框
        draw.rectangle([x, y, x+w, y+h], outline=COLOR_GRID, width=1)
        # 居中文字
        bbox = draw.textbbox((0, 0), text, font=f)
        text_w = bbox[2] - bbox[0]
        text_h = bbox[3] - bbox[1]
        draw.text((x + (w - text_w) / 2, y + (h - text_h) / 2 - 2), text, font=f, fill=COLOR_TEXT)

    # --- 5. 绘制顶部日期 (合并单元格效果) ---
    draw_cell_text(0, 0, img_width, HEADER_H, f"{date_str} 战况表", f=font_bold)

    # --- 6. 绘制表头 (局数) ---
    # 第一列头为空 (或者写"玩家")
    start_y = HEADER_H
    draw_cell_text(0, start_y, CELL_W_NAME, ROW_H, "玩家/局数", f=font_bold)
    
    for i, (mid, ts, _) in enumerate(matches_data):
        x = CELL_W_NAME + i * CELL_W_GAME
        # 显示 "第一局(18:30)" 或简写 "第1局"
        header_text = f"第{i+1}局"
        draw_cell_text(x, start_y, CELL_W_GAME, ROW_H, header_text, f=font_bold)

    # --- 7. 绘制内容 (玩家行) ---
    for r, player in enumerate(sorted_players):
        curr_y = HEADER_H + ROW_H + r * ROW_H
        
        # 7.1 绘制左侧玩家名
        draw_cell_text(0, curr_y, CELL_W_NAME, ROW_H, player, f=font_bold)
        
        # 7.2 绘制该玩家每一局的情况
        for c, (mid, _, _) in enumerate(matches_data):
            curr_x = CELL_W_NAME + c * CELL_W_GAME
            
            p_data = data_map[mid].get(player)
            
            if p_data:
                # 组合文字: "曹操 郭嘉"
                content = f"{p_data['main']}  {p_data['vice']}"
                # 判断背景色: 赢了是黄色，输了是白色(None)
                bg_color = COLOR_WIN_BG if p_data['win'] else None
                
                draw_cell_text(curr_x, curr_y, CELL_W_GAME, ROW_H, content, bg=bg_color)
            else:
                # 没参加这一局，画个斜线或者留白，这里留白
                draw_cell_text(curr_x, curr_y, CELL_W_GAME, ROW_H, "")

    bio = BytesIO()
    image.save(bio, 'PNG')
    bio.seek(0)
    return bio



# ================= 图片生成工具 =================
def draw_text_image(text_lines, title=None, highlight_color=None):
    """生成文字图片"""
    try:
        font = ImageFont.truetype(FONT_PATH, FONT_SIZE)
        title_font = ImageFont.truetype(FONT_PATH, FONT_SIZE + 6)
    except IOError:
        font = ImageFont.load_default()
        title_font = ImageFont.load_default()
    
    # 预计算宽高
    dummy_draw = ImageDraw.Draw(Image.new('RGB', (1, 1)))
    max_width = 0
    total_height = 20
    
    content = text_lines
    if title:
        bbox = dummy_draw.textbbox((0, 0), title, font=title_font)
        max_width = max(max_width, bbox[2] - bbox[0])
        total_height += (bbox[3] - bbox[1]) + 20

    line_height = FONT_SIZE + 10
    for line in content:
        bbox = dummy_draw.textbbox((0, 0), line, font=font)
        max_width = max(max_width, bbox[2] - bbox[0])
        total_height += line_height
            
    img_width = max_width + 60
    img_height = total_height + 40
    
    image = Image.new('RGB', (img_width, img_height), color=(250, 250, 250))
    draw = ImageDraw.Draw(image)
    
    y = 30
    if title:
        draw.text((30, y), title, font=title_font, fill=(0, 0, 0))
        y += line_height + 10
        draw.line((30, y, img_width-30, y), fill=(200, 200, 200), width=2)
        y += 10

    for line in content:
        # 简单着色：如果是分隔线或特殊标记
        fill_color = (0, 0, 0)
        if "=== 修改目标" in line or "修改前" in line or "修改后" in line:
            fill_color = (100, 100, 100)
        
        draw.text((30, y), line, font=font, fill=fill_color)
        y += line_height
        
    bio = BytesIO()
    image.save(bio, 'PNG')
    bio.seek(0)
    return bio

# ================= 逻辑解析 =================
def parse_match_data(text):
    """解析并标准化对局数据"""
    lines = text.strip().split('\n')
    players = []
    
    start_idx = 0
    if lines[0].strip().startswith('#'):
        start_idx = 1
        
    for line in lines[start_idx:]:
        parts = line.strip().replace(',', ' ').replace('，', ' ').replace('：', ' ').replace(':', ' ').split()
        if len(parts) < 3:
            continue
        
        player_name = parts[0]
        raw_main = parts[1]
        
        rest = parts[2:]
        raw_vice = rest[0]
        
        # 判定胜负
        is_winner = 0
        full_line = line.strip()
        # 只要包含 emoji 或者 'win' 字样即视为胜利
        if emoji.emoji_count(full_line) > 0 or 'win' in full_line.lower() or (']' in full_line and '[' in full_line):
             is_winner = 1
        
        # 清理副将名称中的emoji
        clean_vice = ''.join(c for c in raw_vice if c.isalnum() or '\u4e00' <= c <= '\u9fff')
        
        # === 调用 Config 进行名字标准化 ===
        norm_main = config.normalize_name(raw_main)
        norm_vice = config.normalize_name(clean_vice)
        
        players.append({
            'player': player_name,
            'main': norm_main,
            'vice': norm_vice,
            'winner': is_winner
        })
    return players

def format_match_text(match_id, timestamp, players):
    lines = [f"🔢 局号: {match_id}  |  🕒 {timestamp}"]
    lines.append("-" * 30)
    for p in players:
        win_mark = " 💅" if p['is_winner'] else ""
        lines.append(f"{p['player'].ljust(6)} {p['main_gen']} {p['vice_gen']}{win_mark}")
    return lines

# ================= 消息处理 =================

async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    # 1. 处理密码输入的特殊状态 (用于 truncate)
    if context.user_data.get('awaiting_password'):
        if update.message.text == TRUNCATE_PASSWORD:
            conn = get_db_connection()
            c = conn.cursor()
            c.execute("DROP TABLE IF EXISTS details")
            c.execute("DROP TABLE IF EXISTS matches")
            conn.commit()
            conn.close()
            init_db() # 重建空表
            context.user_data['awaiting_password'] = False
            await update.message.reply_text("💥 数据库已清空 (Truncated)。")
        else:
            context.user_data['awaiting_password'] = False
            await update.message.reply_text("❌ 密码错误，操作取消。")
        return

    text = update.message.text
    if not text or not text.startswith('#'):
        return

    # 分割命令
    parts = text.split(maxsplit=1)
    # 兼容 # command 和 #command
    if parts[0] == '#':
        if len(parts) > 1:
            real_parts = parts[1].split(maxsplit=1)
            command = real_parts[0].lower()
            payload = real_parts[1] if len(real_parts) > 1 else ""
        else:
            return
    else:
        command = parts[0][1:].lower()
        payload = parts[1] if len(parts) > 1 else ""

    # --- 功能 1: 加入对局 (#add) ---
    if command == 'add':
        players = parse_match_data(text)
        if not players:
            await update.message.reply_text("❌ 格式错误或未识别到数据。")
            return
        
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M')
        conn = get_db_connection()
        c = conn.cursor()
        c.execute("INSERT INTO matches (timestamp) VALUES (?)", (timestamp,))
        match_id = c.lastrowid
        
        for p in players:
            c.execute("INSERT INTO details (match_id, player, main_gen, vice_gen, is_winner) VALUES (?, ?, ?, ?, ?)",
                      (match_id, p['player'], p['main'], p['vice'], p['winner']))
        conn.commit()
        conn.close()
        await update.message.reply_text(f"✅ 第 {match_id} 局记录成功！")

    # --- 功能 2: 查看对局 (#game) ---
    elif command == 'game':
        arg = payload.strip()
        conn = get_db_connection()
        c = conn.cursor()
        
        row = None
        if arg == 'last':
            c.execute("SELECT * FROM matches ORDER BY id DESC LIMIT 1")
            row = c.fetchone()
        elif arg.isdigit():
            c.execute("SELECT * FROM matches WHERE id = ?", (arg,))
            row = c.fetchone()
            
        if not row:
            conn.close()
            await update.message.reply_text("❌ 未找到对局。")
            return
            
        mid, ts = row
        c.execute("SELECT player, main_gen, vice_gen, is_winner FROM details WHERE match_id = ?", (mid,))
        details = c.fetchall()
        conn.close()
        
        p_list = [{'player': r[0], 'main_gen': r[1], 'vice_gen': r[2], 'is_winner': r[3]} for r in details]
        lines = format_match_text(mid, ts, p_list)
        await update.message.reply_text("\n".join(lines))

    # --- 功能 3: 比赛日赛况 (#date) [图片展示] ---
   # --- 修改功能 3: 比赛日赛况 (#date) ---
    if command == 'date':
        date_str = payload.strip() # YYYY-MM-DD
        if not date_str:
            date_str = datetime.now().strftime('%Y-%m-%d') # 默认当天
            
        conn = get_db_connection()
        c = conn.cursor()
        
        # 1. 查出当天所有局
        c.execute("SELECT id, timestamp FROM matches WHERE timestamp LIKE ? ORDER BY id ASC", (f"{date_str}%",))
        matches_raw = c.fetchall()
        
        if not matches_raw:
            conn.close()
            await update.message.reply_text(f"📅 {date_str} 无比赛记录。")
            return

        # 2. 组装复杂数据结构
        # matches_data = [ (mid, ts, [ (player, main, vice, win), ... ]), ... ]
        matches_data = []
        all_players = set()
        
        for mid, ts in matches_raw:
            c.execute("SELECT player, main_gen, vice_gen, is_winner FROM details WHERE match_id = ?", (mid,))
            details = c.fetchall()
            matches_data.append((mid, ts, details))
            
            # 收集所有出现过的玩家，用于生成行头
            for p, _, _, _ in details:
                all_players.add(p)
            
        conn.close()
        
        # 3. 调用新的Excel绘图函数
        try:
            img_bio = draw_excel_style_image(date_str, matches_data, all_players)
            await update.message.reply_photo(photo=img_bio, caption=f"📅 {date_str} 战况表")
        except Exception as e:
            await update.message.reply_text(f"❌ 生成图片出错: {e}")

    # --- 功能 4: 修改对局 (#update) [Confirm + 图片] ---
    elif command == 'update':
        # payload 应该是 ID \n data
        # 需要重新解析一下 text 拿到 ID
        lines = text.split('\n')
        header = lines[0].split()
        target_id = None
        
        # 尝试提取ID
        for part in header:
            if part.isdigit():
                target_id = part
                break
        
        if not target_id:
            await update.message.reply_text("❌ 请指定局号，例如 #update 10")
            return

        new_players = parse_match_data(text) # 利用现有的解析逻辑
        if not new_players:
            await update.message.reply_text("❌ 请在命令下方输入新的对局数据。")
            return

        conn = get_db_connection()
        c = conn.cursor()
        c.execute("SELECT * FROM matches WHERE id = ?", (target_id,))
        if not c.fetchone():
            conn.close()
            await update.message.reply_text("❌ 局号不存在。")
            return

        # 获取旧数据用于对比
        c.execute("SELECT player, main_gen, vice_gen, is_winner FROM details WHERE match_id = ?", (target_id,))
        old_data = c.fetchall()
        conn.close()

        # 生成对比图文本
        comp_lines = ["[原有数据]:"]
        for r in old_data:
            mark = " 💅" if r[3] else ""
            comp_lines.append(f"{r[0]} {r[1]} {r[2]}{mark}")
        
        comp_lines.append("")
        comp_lines.append("[更新为]:")
        for p in new_players:
            mark = " 💅" if p['winner'] else ""
            comp_lines.append(f"{p['player']} {p['main']} {p['vice']}{mark}")

        img = draw_text_image(comp_lines, title=f"⚠️ 确认更新 第 {target_id} 局")
        
        # 存入上下文
        context.user_data['action'] = 'update'
        context.user_data['tid'] = target_id
        context.user_data['payload'] = new_players

        kb = [[InlineKeyboardButton("✅ 确认更新", callback_data='confirm'),
               InlineKeyboardButton("❌ 取消", callback_data='cancel')]]
        await update.message.reply_photo(img, caption="请确认修改内容：", reply_markup=InlineKeyboardMarkup(kb))

    # --- 功能 5: 删除对局 (#remove) [Confirm] ---
    elif command == 'remove':
        target_id = payload.strip()
        if not target_id.isdigit():
            await update.message.reply_text("❌ 格式错误: #remove ID")
            return
            
        conn = get_db_connection()
        c = conn.cursor()
        c.execute("SELECT timestamp FROM matches WHERE id = ?", (target_id,))
        res = c.fetchone()
        if not res:
            conn.close()
            await update.message.reply_text("❌ 找不到局号。")
            return
            
        c.execute("SELECT player, main_gen, vice_gen, is_winner FROM details WHERE match_id = ?", (target_id,))
        details = c.fetchall()
        conn.close()
        
        info_lines = []
        for r in details:
            mark = " 💅" if r[3] else ""
            info_lines.append(f"{r[0]} {r[1]} {r[2]}{mark}")

        context.user_data['action'] = 'remove'
        context.user_data['tid'] = target_id
        
        kb = [[InlineKeyboardButton("🗑️ 确认删除", callback_data='confirm'),
               InlineKeyboardButton("❌ 取消", callback_data='cancel')]]
        await update.message.reply_text(f"⚠️ 确定删除第 {target_id} 局吗？\n\n" + "\n".join(info_lines), reply_markup=InlineKeyboardMarkup(kb))

    # --- 功能 6: 展示胜率 (#rates) ---
    elif command == 'rates':
        conn = get_db_connection()
        c = conn.cursor()
        c.execute("SELECT player, is_winner FROM details")
        data = c.fetchall()
        conn.close()
        
        stats = defaultdict(lambda: {'win': 0, 'total': 0})
        for player, is_win in data:
            stats[player]['total'] += 1
            if is_win:
                stats[player]['win'] += 1
        
        # 排序：胜率 -> 胜场 -> 总场
        ranking = sorted(stats.items(), 
                         key=lambda x: (x[1]['win']/x[1]['total'], x[1]['win'], x[1]['total']), 
                         reverse=True)
        
        out = ["🏆 玩家胜率榜 🏆", "-"*25, "玩家   | 胜/总 | 胜率"]
        for p, s in ranking:
            rate = (s['win'] / s['total']) * 100
            out.append(f"{p.ljust(6)} | {s['win']}/{s['total']} | {rate:.1f}%")
            
        await update.message.reply_text("\n".join(out))

    # --- 功能 7: 清空数据库 (#truncate) ---
    elif command == 'truncate':
        kb = [[InlineKeyboardButton("⚠️ 我确定要清空", callback_data='req_truncate_pwd'),
               InlineKeyboardButton("❌ 点错了", callback_data='cancel')]]
        await update.message.reply_text("⛔️ 警告：此操作将永久删除所有对局记录！\n如果你确定要继续，请点击下方按钮。", reply_markup=InlineKeyboardMarkup(kb))

    # --- 功能 8: 发送 Excel 记录 (#email) ---
    elif command == 'email':
        target_email = payload.strip()
        # 简单验证邮箱格式
        if '@' not in target_email or '.' not in target_email:
            await update.message.reply_text("❌ 邮箱格式看起来不正确，请检查。")
            return

        await update.message.reply_text("⏳ 正在生成 Excel 并发送，请稍候...")
        
        # 1. 生成文件
        try:
            excel_data = generate_excel_bytes()
            if not excel_data:
                await update.message.reply_text("❌ 数据库是空的，没有记录可发送。")
                return
        except Exception as e:
            await update.message.reply_text(f"❌ 生成 Excel 失败: {e}")
            return

        # 2. 发送邮件 (建议放入线程池或异步任务，但在简单 Bot 中直接调用也行)
        success, msg = send_email_with_excel(target_email, excel_data)
        
        if success:
            await update.message.reply_text(f"✅ 邮件已发送至 {target_email}，请查收（可能在垃圾箱）。")
        else:
            await update.message.reply_text(f"❌ 发送失败: {msg}\n请检查 config.py 中的邮箱配置。")

# ================= 按钮回调 =================

async def button_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    data = query.data
    
    if data == 'cancel':
        await query.edit_message_caption("❌ 操作已取消") if query.message.caption else await query.edit_message_text("❌ 操作已取消")
        context.user_data.clear()
        return

    # Truncate 流程中间步
    if data == 'req_truncate_pwd':
        context.user_data['awaiting_password'] = True
        await query.edit_message_text("🔒 请回复密码以执行清空操作：")
        return

    # Update / Remove 确认流程
    action = context.user_data.get('action')
    tid = context.user_data.get('tid')
    
    conn = get_db_connection()
    c = conn.cursor()
    
    try:
        if action == 'update' and data == 'confirm':
            new_data = context.user_data.get('payload')
            c.execute("DELETE FROM details WHERE match_id = ?", (tid,))
            for p in new_data:
                c.execute("INSERT INTO details (match_id, player, main_gen, vice_gen, is_winner) VALUES (?, ?, ?, ?, ?)",
                          (tid, p['player'], p['main'], p['vice'], p['winner']))
            conn.commit()
            await query.edit_message_caption(f"✅ 第 {tid} 局更新成功。")
            
        elif action == 'remove' and data == 'confirm':
            c.execute("DELETE FROM details WHERE match_id = ?", (tid,))
            c.execute("DELETE FROM matches WHERE id = ?", (tid,))
            conn.commit()
            await query.edit_message_text(f"✅ 第 {tid} 局已删除。")
            
    except Exception as e:
        await query.message.reply_text(f"❌ 数据库错误: {e}")
    finally:
        conn.close()
        context.user_data.clear()

# ================= 启动 =================
if __name__ == '__main__':
    init_db()
    app = ApplicationBuilder().token(BOT_TOKEN).build()
    
    # 过滤所有文本消息，交给 handle_message 处理逻辑判断
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
    app.add_handler(CallbackQueryHandler(button_callback))
    
    print("Bot is running...")
    app.run_polling()